"""
진단 결과 리포트 생성기 (Excel / Markdown / Log)
"""
import json
import os
from collections import OrderedDict
from datetime import datetime
from core.result import ScanReport, Status, Severity

import sys as _sys
if getattr(_sys, "frozen", False):
    # PyInstaller 실행 파일 → 바탕화면에 저장
    REPORTS_DIR = os.path.join(os.path.expanduser("~"), "Desktop", "vuln-scanner-reports")
else:
    # 개발 환경 → 프로젝트 루트 reports/
    REPORTS_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "reports")

_SECTION_NAMES: dict[str, str] = {
    "1": "계정 관리",
    "2": "파일 시스템",
    "3": "네트워크 서비스",
    "4": "로그 관리",
    "5": "주요 응용 설정",
    "6": "시스템 보안 설정",
    "7": "보안 패치",
}

_STATUS_FILL = {
    "취약":    "FFEB9C9C",
    "양호":    "FFC6EFCE",
    "수동점검": "FFFFEB9C",
    "오류":    "FFD9D9D9",
    "미해당":  "FFEDEDED",
}
_SEVERITY_COLOR = {
    "위험": "FFC00000",  # 진한 빨강
    "높음": "FFFF0000",  # 빨강
    "보통": "FFED7D31",  # 주황
    "낮음": "FF7F7F7F",  # 회색
    "정보": "FF4472C4",  # 파랑
}
_STATUS_LABEL = {
    "취약":    "[취약]",
    "양호":    "[양호]",
    "수동점검": "[검토]",
    "오류":    "[검토]",
    "미해당":  "[ N/A]",
}
_STATUS_ICON = {
    "취약":    "🔴",
    "양호":    "🟢",
    "수동점검": "🟡",
    "오류":    "⚫",
    "미해당":  "⚪",
}
_SEVERITY_ICON = {
    "위험": "🚨",
    "높음": "🔴",
    "보통": "🟠",
    "낮음": "🟡",
    "정보": "🔵",
}


def _ensure_dir():
    os.makedirs(REPORTS_DIR, exist_ok=True)


def _filename(report: ScanReport, ext: str) -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_type = report.scan_type.replace("/", "_").replace(" ", "_")
    return os.path.join(REPORTS_DIR, f"{safe_type}_{ts}.{ext}")


def _sections(report: ScanReport) -> dict[str, list]:
    secs: dict[str, list] = OrderedDict()
    for r in report.results:
        key = r.check_id.split(".")[0] if "." in r.check_id else "기타"
        secs.setdefault(key, []).append(r)
    return secs


# ══════════════════════════════════════════════════════════════════════
# Cmd-Log  ── 명령어 + 실행 결과만 (자동 저장용 간결 로그)
# ══════════════════════════════════════════════════════════════════════

def save_cmd_log(report: ScanReport) -> str:
    """명령어와 실행 결과 로우 데이터만 저장 (자동 저장)."""
    _ensure_dir()
    path  = _filename(report, "log")
    lines: list[str] = []

    for r in report.results:
        if not r.command and not r.cmd_output:
            continue
        if r.command:
            lines.append(r.command.strip())
        if r.cmd_output:
            lines.append(r.cmd_output.strip())
        lines.append("")

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    return path


# ══════════════════════════════════════════════════════════════════════
# Log  ── 명령어 + 결과값 + 판정 모두 포함 (TXT 버튼 저장용)
# ══════════════════════════════════════════════════════════════════════

def save_log(report: ScanReport) -> str:
    """섹션별·항목별로 진단 명령어·결과값·판정을 포함하는 상세 텍스트 저장."""
    _ensure_dir()
    path = _filename(report, "txt")
    s = report.summary
    W = 70

    lines: list[str] = []

    def hr(c="═"):
        lines.append(c * W)

    def add(text: str = ""):
        lines.append(text)

    # ── 헤더 ──────────────────────────────────────────────────────────
    hr()
    add(f"  취약점 진단 로그  |  {report.scan_type}")
    hr()
    add(f"  대상  : {report.target}")
    add(f"  시작  : {report.started_at}")
    add(f"  종료  : {report.finished_at or '-'}")
    hr("─")
    review = s["manual"] + s["error"]
    add(f"  전체 {s['total']}건  "
        f"[취약] {s['vulnerable']}  "
        f"[양호] {s['safe']}  "
        f"[검토] {review}  "
        f"[N/A] {s['skipped']}")
    add(f"  위험도  위험:{s['by_severity']['위험']}  "
        f"높음:{s['by_severity']['높음']}  "
        f"보통:{s['by_severity']['보통']}  "
        f"낮음:{s['by_severity']['낮음']}")
    hr()
    add()

    # ── 섹션별 상세 ───────────────────────────────────────────────────
    for sec_key, results in _sections(report).items():
        sec_name = _SECTION_NAMES.get(sec_key, f"섹션 {sec_key}")
        hr("─")
        add(f"  {sec_key}. {sec_name}")
        hr("─")
        add()

        for r in results:
            label = _STATUS_LABEL.get(r.status.value, "[    ]")

            # ① 항목 헤더
            add(f"{'─' * 60}")
            add(f"{label} [{r.check_id}]  {r.name}")
            add(f"{'─' * 60}")

            # ② 위험도 / 설명
            add(f"  위험도    : {r.severity.value}")
            add(f"  점검 내용 : {r.description}")
            add()

            # ③ 진단 명령어 (있는 경우만)
            if r.command:
                add("  [ 진단 명령어 ]")
                for cmd_line in r.command.strip().splitlines():
                    add(f"    $ {cmd_line}")
                add()

            # ④ 명령어 실행 결과 (있는 경우만)
            if r.cmd_output:
                add("  [ 명령어 실행 결과 ]")
                out_lines = r.cmd_output.strip().splitlines()
                for ol in out_lines[:30]:          # 최대 30줄
                    add(f"    {ol}")
                if len(out_lines) > 30:
                    add(f"    ... (이하 {len(out_lines) - 30}줄 생략)")
                add()
            elif r.command:
                add("  [ 명령어 실행 결과 ]")
                add("    (출력 없음)")
                add()

            # ⑤ 판정 결과
            add(f"  [ 판정 결과 : {label} ]")
            for detail_line in r.details.strip().splitlines():
                add(f"    {detail_line}")
            add()

            # ⑥ 조치 방안
            add("  [ 조치 방안 ]")
            for rec_line in r.recommendation.strip().splitlines():
                add(f"    {rec_line}")
            add()

    # ── 푸터 ──────────────────────────────────────────────────────────
    hr()
    add(f"  로그 생성: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    hr()

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    return path


# ══════════════════════════════════════════════════════════════════════
# Excel  ── 판정 결과 + 명령어·결과값 컬럼 포함
# ══════════════════════════════════════════════════════════════════════

def save_excel(report: ScanReport) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    _ensure_dir()
    path = _filename(report, "xlsx")
    wb = Workbook()

    hdr_font  = Font(name="맑은 고딕", bold=True, size=10, color="FFFFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="FF2C3E50")
    val_font  = Font(name="맑은 고딕", size=9)
    bold_font = Font(name="맑은 고딕", bold=True, size=10)
    center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left      = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
    thin      = Side(style="thin", color="FFBDC3C7")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    s = report.summary

    # ── 시트 1: 요약 ──────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "요약"
    ws_sum["A1"] = "취약점 진단 결과 보고서"
    ws_sum["A1"].font = Font(name="맑은 고딕", bold=True, size=14)
    ws_sum.merge_cells("A1:C1")
    ws_sum.row_dimensions[1].height = 30

    for i, (label, value) in enumerate([
        ("진단 유형", report.scan_type),
        ("진단 대상", report.target),
        ("시작 시각", report.started_at),
        ("종료 시각", report.finished_at or "-"),
    ], start=3):
        ws_sum.cell(i, 1, label).font = Font(name="맑은 고딕", bold=True, size=10)
        ws_sum.cell(i, 2, value).font = val_font
        ws_sum.row_dimensions[i].height = 18

    ws_sum.cell(8, 1, "구분").font  = hdr_font
    ws_sum.cell(8, 2, "건수").font  = hdr_font
    ws_sum.cell(8, 1).fill  = hdr_fill
    ws_sum.cell(8, 2).fill  = hdr_fill
    ws_sum.cell(8, 1).alignment = center
    ws_sum.cell(8, 2).alignment = center

    review = s["manual"] + s["error"]
    stat_rows = [
        ("전체",    s["total"],               "FFD9EAD3", None),
        ("[취약]",  s["vulnerable"],          "FFEB9C9C", None),
        ("[양호]",  s["safe"],                "FFC6EFCE", None),
        ("[검토]",  review,                   "FFFFEB9C", None),
        ("[ N/A]",  s["skipped"],             "FFEDEDED", None),
        ("",        "",                       "",         None),
        ("위험",    s["by_severity"]["위험"], None, "FFC00000"),
        ("높음",    s["by_severity"]["높음"], None, "FFFF0000"),
        ("보통",    s["by_severity"]["보통"], None, "FFED7D31"),
        ("낮음",    s["by_severity"]["낮음"], None, "FF7F7F7F"),
    ]
    for i, (lbl, val, color, txt_color) in enumerate(stat_rows, start=9):
        if not lbl:
            continue
        cl = ws_sum.cell(i, 1, lbl)
        cv = ws_sum.cell(i, 2, val)
        fc = txt_color or "FF000000"
        cl.font = Font(name="맑은 고딕", bold=True, size=10, color=fc)
        cv.font = Font(name="맑은 고딕", bold=True, size=12, color=fc)
        cl.alignment = cv.alignment = center
        if color:
            fill = PatternFill("solid", fgColor=color)
            cl.fill = cv.fill = fill
        cl.border = cv.border = border
        ws_sum.row_dimensions[i].height = 22

    ws_sum.column_dimensions["A"].width = 12
    ws_sum.column_dimensions["B"].width = 10
    ws_sum.column_dimensions["C"].width = 35

    # ── 시트 2: 전체 결과 ─────────────────────────────────────────────
    COLS = [
        ("ID",        8),  ("카테고리",   14), ("점검 항목",  26),
        ("결과",       9),  ("위험도",     9),  ("점검 설명",  35),
        ("진단 결과",  35), ("조치 방안",  35),
        ("진단 명령어", 40), ("명령어 출력", 40), ("점검 시각",  18),
    ]

    def _make_sheet(wb, title: str, fill_color: str, rows):
        ws = wb.create_sheet(title)
        for ci, (col, width) in enumerate(COLS, start=1):
            c = ws.cell(1, ci, col)
            c.font      = hdr_font
            c.fill      = PatternFill("solid", fgColor=fill_color)
            c.alignment = center
            c.border    = border
            ws.column_dimensions[get_column_letter(ci)].width = width
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

        for ri, r in enumerate(rows, start=2):
            is_vuln = r.status == Status.VULNERABLE
            row_font = Font(name="맑은 고딕", bold=is_vuln, size=9)
            data = [
                r.check_id, r.category, r.name,
                r.status.value, r.severity.value,
                r.description, r.details, r.recommendation,
                r.command, r.cmd_output, r.checked_at,
            ]
            for ci, val in enumerate(data, start=1):
                cell = ws.cell(ri, ci, val)
                cell.font      = row_font
                cell.alignment = left
                cell.border    = border
            # 결과 컬럼: 배경색 유지
            ws.cell(ri, 4).fill      = PatternFill("solid", fgColor=_STATUS_FILL.get(r.status.value, "FFFFFFFF"))
            ws.cell(ri, 4).alignment = center
            ws.cell(ri, 4).font      = Font(name="맑은 고딕", bold=is_vuln, size=9)
            # 위험도 컬럼: 배경 없이 텍스트 색상만
            sev_color = _SEVERITY_COLOR.get(r.severity.value, "FF000000")
            ws.cell(ri, 5).font      = Font(name="맑은 고딕", bold=True, size=9, color=sev_color)
            ws.cell(ri, 5).alignment = center
            ws.row_dimensions[ri].height = 60
        return ws

    _make_sheet(wb, "전체 결과", "FF2C3E50", report.results)
    _make_sheet(wb, "취약 항목", "FFC0392B",
                [r for r in report.results if r.status == Status.VULNERABLE])

    wb.save(path)
    return path


# ══════════════════════════════════════════════════════════════════════
# Markdown  ── 섹션별 정리 + 진단 명령어 코드블록
# ══════════════════════════════════════════════════════════════════════

def save_markdown(report: ScanReport) -> str:
    _ensure_dir()
    path = _filename(report, "md")
    s = report.summary
    lines: list[str] = []

    review = s["manual"] + s["error"]

    # 헤더
    lines += [
        f"# 취약점 진단 결과 — {report.scan_type}",
        "",
        f"| 항목 | 내용 |",
        f"|------|------|",
        f"| 진단 대상 | `{report.target}` |",
        f"| 시작 시각 | {report.started_at} |",
        f"| 종료 시각 | {report.finished_at or '-'} |",
        "",
        "## 요약",
        "",
        "| 구분 | 건수 |",
        "|------|-----:|",
        f"| 전체        | **{s['total']}** |",
        f"| 🔴 **취약** | **{s['vulnerable']}** |",
        f"| 🟢 양호     | {s['safe']} |",
        f"| 🟡 검토     | {review} |",
        f"| ⚪ N/A      | {s['skipped']} |",
        "",
        "| 위험도 | 건수 |",
        "|--------|-----:|",
        f"| 🚨 위험 | **{s['by_severity']['위험']}** |",
        f"| 🔴 높음 | **{s['by_severity']['높음']}** |",
        f"| 🟠 보통 | {s['by_severity']['보통']} |",
        f"| 🟡 낮음 | {s['by_severity']['낮음']} |",
        "",
    ]

    # 섹션별 상세
    for sec_key, results in _sections(report).items():
        sec_name = _SECTION_NAMES.get(sec_key, f"섹션 {sec_key}")
        lines += [f"---", f"## {sec_key}. {sec_name}", ""]

        for r in results:
            st_icon  = _STATUS_ICON.get(r.status.value, "")
            sev_icon = _SEVERITY_ICON.get(r.severity.value, "")
            label    = _STATUS_LABEL.get(r.status.value, "")

            lines += [
                f"### {label} [{r.check_id}] {r.name}",
                "",
                f"| 항목 | 내용 |",
                f"|------|------|",
                f"| 판정     | {st_icon} **{r.status.value}** |",
                f"| 위험도   | {sev_icon} {r.severity.value} |",
                f"| 점검 내용 | {r.description} |",
                f"| 진단 결과 | {r.details.replace(chr(10), ' / ')} |",
                f"| 조치 방안 | {r.recommendation.replace(chr(10), ' / ')} |",
            ]

            if r.command:
                lines += [
                    "",
                    "**진단 명령어**",
                    "```bash",
                    r.command.strip(),
                    "```",
                ]

            if r.cmd_output:
                out_lines = r.cmd_output.strip().splitlines()
                preview   = "\n".join(out_lines[:20])
                suffix    = f"\n... ({len(out_lines) - 20}줄 생략)" if len(out_lines) > 20 else ""
                lines += [
                    "",
                    "**명령어 실행 결과**",
                    "```",
                    preview + suffix,
                    "```",
                ]
            elif r.command:
                lines += ["", "**명령어 실행 결과**", "```", "(출력 없음)", "```"]

            lines.append("")

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    return path


# ══════════════════════════════════════════════════════════════════════
# JSON (내부 저장용)
# ══════════════════════════════════════════════════════════════════════

def save_json(report: ScanReport) -> str:
    _ensure_dir()
    path = _filename(report, "json")
    data = {
        "target": report.target,
        "scan_type": report.scan_type,
        "started_at": report.started_at,
        "finished_at": report.finished_at,
        "summary": report.summary,
        "results": [r.to_dict() for r in report.results],
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return path
