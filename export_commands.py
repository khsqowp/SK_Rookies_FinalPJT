"""
보안 진단 스크립트 — 항목별 명령어 추출 도구
GUI에서 호출하거나 단독 실행 가능.
"""
import sys

# ══════════════════════════════════════════════════════════════════
# SK 표준 가이드 명령어
# ══════════════════════════════════════════════════════════════════

_SK_LINUX = [
    ("U-01", "root 계정 원격 접속 제한",
     "cat /etc/ssh/sshd_config | grep PermitRootLogin\ncat /etc/securetty"),
    ("U-02", "패스워드 복잡도 설정",
     "cat /etc/pam.d/common-password\ncat /etc/pam.d/system-auth"),
    ("U-03", "계정 잠금 임계값 설정",
     "cat /etc/pam.d/common-auth\ncat /etc/pam.d/system-auth"),
    ("U-04", "패스워드 파일 보호",
     "ls -al /etc/passwd\nls -al /etc/shadow"),
    ("U-05", "root 이외 UID 0 금지",
     "awk -F: '$3 == 0 { print $0 }' /etc/passwd"),
    ("U-06", "root 계정 su 제한",
     "cat /etc/pam.d/su\ncat /etc/group | grep wheel"),
    ("U-07", "패스워드 최소 길이 설정",
     "grep PASS_MIN_LEN /etc/login.defs"),
    ("U-08", "패스워드 최대 사용 기간 설정",
     "grep PASS_MAX_DAYS /etc/login.defs"),
    ("U-09", "패스워드 최소 사용 기간 설정",
     "grep PASS_MIN_DAYS /etc/login.defs"),
    ("U-10", "불필요한 계정 제거",
     'egrep "^(lp|uucp|nuucp):" /etc/passwd | egrep -v "false|nologin"'),
    ("U-11", "관리자 그룹 최소 계정 포함",
     "cat /etc/group | grep -E '^wheel|^sudo'"),
    ("U-12", "계정 없는 GID 금지",
     "cat /etc/passwd\ncat /etc/group"),
    ("U-13", "동일한 UID 금지",
     "awk -F: '{print $3}' /etc/passwd | sort | uniq -d"),
    ("U-14", "사용자 shell 점검",
     r'grep -v "nologin\|false" /etc/passwd | awk -F: \'$7!="" {print $1,$7}\''),
    ("U-15", "세션 타임아웃 설정",
     "cat /etc/profile | grep TMOUT\ncat /etc/bashrc | grep TMOUT"),
    ("U-16", "홈 디렉터리 소유자·권한 설정",
     "ls -al /home"),
    ("U-18", "world writable 파일 점검",
     "find / -type f -perm -o+w -exec ls -l {} \\; 2>/dev/null"),
    ("U-19", "/dev 불필요 파일 존재",
     "find /dev -type f -exec ls -l {} \\; 2>/dev/null"),
    ("U-20", "SetUID/SetGID 파일 점검",
     r"find / -user root -type f \( -perm -4000 -o -perm -2000 \) -exec ls -al {} \; 2>/dev/null"),
    ("U-22", "/etc/passwd 소유자·권한",
     "ls -al /etc/passwd"),
    ("U-23", "/etc/shadow 소유자·권한",
     "ls -al /etc/shadow"),
    ("U-24", "/etc/hosts 소유자·권한",
     "ls -al /etc/hosts"),
    ("U-27", "/etc/services 소유자·권한",
     "ls -al /etc/services"),
    ("U-29", "최신 보안 패치 적용",
     "uname -a\ncat /etc/os-release"),
    ("U-30", "ftpusers 파일 소유자·권한",
     "ls -al /etc/ftpusers 2>/dev/null"),
    ("U-31", "ftp root 계정 제한",
     "cat /etc/ftpusers 2>/dev/null"),
    ("U-32", "at 파일 소유자·권한",
     "ls -al /etc/at.allow 2>/dev/null\nls -al /etc/at.deny 2>/dev/null"),
    ("U-33", "cron 파일 소유자·권한",
     "ls -al /etc/crontab 2>/dev/null\nls -al /etc/cron.d/ 2>/dev/null"),
    ("U-34", "hosts.equiv 파일 제거",
     "ls -al /etc/hosts.equiv 2>/dev/null"),
    ("U-35", "익명 FTP 제한",
     "cat /etc/vsftpd.conf 2>/dev/null | grep anonymous"),
    ("U-36", "r계열 서비스 비활성화",
     "systemctl is-active rsh rlogin rexec 2>/dev/null"),
    ("U-37", "NFS 설정 파일 점검",
     "cat /etc/exports 2>/dev/null"),
    ("U-42", "tftp, talk 서비스 비활성화",
     "systemctl is-active tftp talk 2>/dev/null"),
    ("U-46", "DNS 보안 버전 패치",
     "named -v 2>/dev/null"),
    ("U-47", "DNS Zone Transfer 설정",
     "cat /etc/named.conf 2>/dev/null | grep allow-transfer"),
    ("U-48", "Apache 디렉터리 리스팅 제거",
     "cat /etc/apache2/apache2.conf 2>/dev/null | grep -i options"),
    ("U-56", "Apache 불필요한 모듈 제거",
     "apache2 -M 2>/dev/null\nhttpd -M 2>/dev/null"),
    ("U-60", "SNMP Community String 복잡성",
     "cat /etc/snmp/snmpd.conf 2>/dev/null | grep -E '^rocommunity|^rwcommunity'"),
    ("U-62", "로그온 배너 설정",
     "cat /etc/issue 2>/dev/null\ncat /etc/issue.net 2>/dev/null"),
    ("U-64", "syslog 로그 설정",
     "cat /etc/rsyslog.conf 2>/dev/null"),
]

_SK_WINDOWS = [
    ("W-1.1", "로컬 계정 사용 설정",
     "net user\nnet user Administrator 2>nul\nnet user guest 2>nul"),
    ("W-1.2", "계정 잠금 정책 설정",
     "net accounts"),
    ("W-1.3", "암호 정책 설정",
     'net accounts\ncmd /c "secedit /export /cfg C:\\Windows\\Temp\\wss.cfg /quiet && type C:\\Windows\\Temp\\wss.cfg"'),
    ("W-1.5", "UAC 설정",
     'reg query "HKLM\\SOFTWARE\\Microsoft\\Windows\\CurrentVersion\\Policies\\System" /v EnableLUA'),
    ("W-1.6", "익명 SID/이름 변환 허용 정책",
     'reg query "HKLM\\SYSTEM\\CurrentControlSet\\Control\\Lsa" /v TurnOffAnonymousBlock'),
    ("W-1.7", "콘솔 로그온 빈 암호 사용 제한",
     'reg query "HKLM\\SYSTEM\\CurrentControlSet\\Control\\Lsa" /v LimitBlankPasswordUse'),
    ("W-1.8", "관리자 그룹 최소 사용자 포함",
     "net localgroup administrators"),
    ("W-2.1", "CMD.EXE 파일 권한 설정",
     "sc query W3SVC\nicacls C:\\Windows\\System32\\cmd.exe"),
    ("W-2.2", "사용자 홈 디렉터리 접근 제한",
     "icacls C:\\Users"),
    ("W-2.3", "공유 폴더 설정",
     'net share\nreg query "HKLM\\SYSTEM\\CurrentControlSet\\Services\\LanmanServer\\Parameters" /v AutoShareServer'),
    ("W-2.4", "SAM 파일 권한 설정",
     "icacls C:\\Windows\\System32\\config\\SAM"),
    ("W-3.1", "불필요한 서비스 제거",
     "sc query type= all state= all"),
    ("W-3.2", "터미널 서비스 암호화 수준",
     'reg query "HKLM\\SYSTEM\\CurrentControlSet\\Control\\Terminal Server\\WinStations\\RDP-Tcp" /v MinEncryptionLevel'),
    ("W-3.3", "NetBIOS 서비스 보안 설정",
     'reg query "HKLM\\SYSTEM\\CurrentControlSet\\Services\\NetBT\\Parameters" /v NetbiosOptions'),
    ("W-3.4", "터미널 서비스 Time Out 설정",
     'reg query "HKLM\\SOFTWARE\\Policies\\Microsoft\\Windows NT\\Terminal Services" /v MaxIdleTime'),
    ("W-4.1", "Telnet 서비스 보안 설정",
     "sc query TlntSvr 2>nul"),
    ("W-4.2", "DNS 보안 설정",
     "sc query DNS"),
    ("W-4.3", "SNMP 서비스 보안 설정",
     'sc query SNMP\nreg query "HKLM\\SYSTEM\\CurrentControlSet\\Services\\SNMP\\Parameters\\ValidCommunities"'),
    ("W-5.1", "원격 로그파일 접근 진단",
     "icacls C:\\Windows\\System32\\config\nicacls C:\\Windows\\System32\\LogFiles"),
    ("W-5.2", "화면 보호기 설정",
     'reg query "HKCU\\Control Panel\\Desktop" /v ScreenSaveActive\nreg query "HKCU\\Control Panel\\Desktop" /v ScreenSaverIsSecure'),
    ("W-5.3", "이벤트 뷰어 설정",
     "wevtutil gl Security\nwevtutil gl System\nwevtutil gl Application"),
    ("W-5.4", "로그인 경고 메시지 표시",
     'reg query "HKLM\\SOFTWARE\\Microsoft\\Windows\\CurrentVersion\\Policies\\System" /v LegalNoticeCaption'),
    ("W-5.5", "마지막 로그온 사용자 계정 숨김",
     'reg query "HKLM\\SOFTWARE\\Microsoft\\Windows\\CurrentVersion\\Policies\\System" /v DontDisplayLastUserName'),
    ("W-5.6", "로그온 없이 시스템 종료 방지",
     'reg query "HKLM\\SOFTWARE\\Microsoft\\Windows\\CurrentVersion\\Policies\\System" /v ShutdownWithoutLogon'),
    ("W-5.7", "로컬 감사 정책 설정",
     "auditpol /get /category:*"),
    ("W-5.8", "가상 메모리 페이지 파일 삭제",
     'reg query "HKLM\\SYSTEM\\CurrentControlSet\\Control\\Session Manager\\Memory Management" /v ClearPageFileAtShutdown'),
    ("W-5.9", "LAN Manager 인증 수준",
     'reg query "HKLM\\SYSTEM\\CurrentControlSet\\Control\\Lsa" /v LmCompatibilityLevel'),
    ("W-5.10", "Everyone 익명 사용자 적용 안함",
     'reg query "HKLM\\SYSTEM\\CurrentControlSet\\Control\\Lsa" /v EveryoneIncludesAnonymous'),
    ("W-5.12", "SMB 세션 유휴 시간 설정",
     'reg query "HKLM\\SYSTEM\\CurrentControlSet\\Services\\LanmanServer\\Parameters" /v AutoDisconnect'),
    ("W-5.13", "예약된 작업 점검",
     'powershell -NoProfile -NonInteractive -Command "Get-ScheduledTask | Where-Object {$_.State -ne \'Disabled\'} | Select TaskName,TaskPath,State"'),
    ("W-5.14", "원격 시스템 종료 권한 설정",
     'powershell -NoProfile -NonInteractive -Command "(Get-Acl -Path \'HKLM:\\SYSTEM\\CurrentControlSet\\Control\\SecurePipeServers\\winreg\').AccessToString"'),
    ("W-5.15", "보안 감사 로그 불가 시 종료 방지",
     'reg query "HKLM\\SYSTEM\\CurrentControlSet\\Control\\Lsa" /v CrashOnAuditFail'),
    ("W-5.16", "보안 채널 데이터 암호화/서명",
     'reg query "HKLM\\SYSTEM\\CurrentControlSet\\Services\\Netlogon\\Parameters" /v RequireSignOrSeal'),
    ("W-6.1", "백신 프로그램 설치",
     'powershell -NoProfile -NonInteractive -Command "Get-MpComputerStatus | Select AMRunningMode,AntivirusEnabled"'),
    ("W-6.2", "최신 엔진 업데이트",
     'powershell -NoProfile -NonInteractive -Command "Get-MpComputerStatus | Select AntivirusSignatureLastUpdated,AntivirusSignatureVersion"'),
    ("W-7.1", "SAM 보안 감사 설정",
     'reg query "HKLM\\SAM" /v Security 2>nul'),
    ("W-7.2", "Null Session 설정",
     'reg query "HKLM\\SYSTEM\\CurrentControlSet\\Control\\Lsa" /v RestrictAnonymous'),
    ("W-7.3", "Remote Registry Service 설정",
     "sc query RemoteRegistry"),
    ("W-7.5", "AutoLogon 제한 설정",
     'reg query "HKLM\\SOFTWARE\\Microsoft\\Windows NT\\CurrentVersion\\Winlogon" /v AutoAdminLogon'),
    ("W-8.1", "최신 서비스 팩 적용",
     'powershell -NoProfile -NonInteractive -Command "[System.Environment]::OSVersion.Version"'),
    ("W-8.2", "최신 HOT FIX 적용",
     'powershell -NoProfile -NonInteractive -Command "Get-HotFix | Sort-Object InstalledOn -Descending | Select-Object -First 10"'),
    ("W-9.1", "OpenSSL 취약점",
     "openssl version 2>nul"),
]

_SK_DBMS = [
    ("DB-01", "DBMS 포트 외부 노출 여부",
     "# MySQL\nnc -zv <host> 3306\n# PostgreSQL\nnc -zv <host> 5432\n# MSSQL\nnc -zv <host> 1433"),
    ("DB-02", "불필요 계정 제거",
     "# MySQL\nSELECT user, host, authentication_string FROM mysql.user;\n# PostgreSQL\nSELECT usename, passwd FROM pg_shadow;\n# MSSQL\nSELECT name, is_disabled FROM sys.server_principals WHERE type = 'S';"),
    ("DB-03", "비밀번호 사용기간·복잡도 설정",
     "# MySQL\nSHOW VARIABLES LIKE 'validate_password%';\n# MSSQL\nSELECT name, is_policy_checked FROM sys.sql_logins;"),
    ("DB-04", "관리자 권한 최소화",
     "# MySQL\nSELECT user, host FROM mysql.user WHERE user='root' AND host != 'localhost';\n# MSSQL\nSELECT m.name FROM sys.server_role_members rm JOIN sys.server_principals m ON rm.member_principal_id=m.principal_id JOIN sys.server_principals r ON rm.role_principal_id=r.principal_id WHERE r.name='sysadmin';"),
    ("DB-05", "감사 로깅 설정",
     "# MySQL\nSHOW VARIABLES LIKE 'general_log%';\n# PostgreSQL\nSHOW log_statement;\n# MSSQL\nSELECT * FROM sys.server_audits;"),
    ("DB-06", "데이터 디렉토리 권한",
     "# MySQL\nSHOW VARIABLES LIKE 'datadir';\n# PostgreSQL\nSHOW data_directory;"),
    ("DB-09", "계정 잠금 정책 설정",
     "# MySQL\nSHOW VARIABLES LIKE 'connection_control%';\n# MSSQL\nSELECT name, is_policy_checked FROM sys.sql_logins;"),
]

_SK_ORACLE = [
    ("OA-01", "불필요 계정 제거",
     "SELECT username, account_status FROM dba_users ORDER BY username;"),
    ("OA-02", "비밀번호 복잡도·사용기간 설정",
     "SELECT resource_name, limit FROM dba_profiles WHERE profile='DEFAULT' AND resource_type='PASSWORD';"),
    ("OA-03", "관리자 권한 최소화",
     "SELECT grantee, granted_role FROM dba_role_privs WHERE granted_role IN ('DBA','SYSDBA','SYSOPER');"),
    ("OA-04", "감사 로깅 설정",
     "SELECT value FROM v$parameter WHERE name='audit_trail';"),
    ("OA-05", "불필요한 PL/SQL 패키지 권한 제한",
     "SELECT grantee, table_name, privilege FROM dba_tab_privs WHERE table_name IN ('UTL_SMTP','UTL_TCP','UTL_HTTP','UTL_FILE','DBMS_SQL') AND grantee='PUBLIC';"),
    ("OA-06", "DB 링크 관리",
     "SELECT owner, db_link, username, host FROM dba_db_links;"),
    ("OA-07", "리스너 비밀번호 설정",
     "lsnrctl status\ncat $ORACLE_HOME/network/admin/listener.ora"),
    ("OA-08", "데이터 디렉토리 권한",
     "SELECT value FROM v$parameter WHERE name='db_create_file_dest';"),
]

_SK_AWS = [
    ("C-1.1", "IAM 사용자 계정 관리",
     "aws iam list-users\naws iam list-attached-user-policies --user-name <USER>"),
    ("C-1.2", "IAM 사용자 계정 단일화",
     "aws iam list-users"),
    ("C-1.3", "IAM 사용자 계정 식별",
     "aws iam list-user-tags --user-name <USER>"),
    ("C-1.4", "IAM 그룹 사용자 관리",
     "aws iam list-groups\naws iam get-group --group-name <GROUP>"),
    ("C-1.5", "EC2 Key Pair 접근 관리",
     "aws ec2 describe-key-pairs"),
    ("C-1.6", "Key Pair 보관 관리",
     "# 수동 점검 — Key Pair 보관 위치 직접 확인"),
    ("C-1.7", "Admin Console 관리자 정책 관리",
     "aws iam list-attached-user-policies --user-name <USER>"),
    ("C-1.8", "Admin Console Access Key 관리",
     "aws iam list-access-keys --user-name <USER>"),
    ("C-1.9", "MFA 설정",
     "aws iam get-account-summary\naws iam list-mfa-devices --user-name <USER>"),
    ("C-1.10", "AWS 계정 패스워드 정책",
     "aws iam get-account-password-policy"),
    ("C-2.1", "인스턴스 서비스 정책 관리",
     "# 수동 점검\naws iam list-roles"),
    ("C-2.2", "네트워크 서비스 정책 관리",
     "# 수동 점검\naws iam list-policies --scope Local"),
    ("C-2.3", "기타 서비스 정책 관리",
     "# 수동 점검\naws iam list-policies --scope Local"),
    ("C-3.1", "보안 그룹 any 포트 허용",
     "aws ec2 describe-security-groups"),
    ("C-3.2", "보안 그룹 any 소스 허용",
     "aws ec2 describe-security-groups"),
    ("C-3.3", "NACL 설정",
     "aws ec2 describe-network-acls"),
    ("C-3.4", "라우팅 테이블 설정",
     "aws ec2 describe-route-tables"),
    ("C-3.5", "인터넷 게이트웨이 설정",
     "aws ec2 describe-internet-gateways"),
    ("C-3.6", "NAT 게이트웨이 설정",
     "aws ec2 describe-nat-gateways"),
    ("C-3.7", "S3 버킷 퍼블릭 접근",
     "aws s3api list-buckets\naws s3api get-public-access-block --bucket <BUCKET>"),
    ("C-3.8", "RDS 서브넷 설정",
     "aws rds describe-db-subnet-groups"),
    ("C-4.1", "EBS 암호화 설정",
     "aws ec2 describe-volumes"),
    ("C-4.2", "RDS 암호화 설정",
     "aws rds describe-db-instances"),
    ("C-4.3", "S3 암호화 설정",
     "aws s3api get-bucket-encryption --bucket <BUCKET>"),
    ("C-4.4", "통신구간 암호화 설정",
     "# 수동 점검 — TLS/SSL, VPN, SSH 적용 여부 확인"),
    ("C-4.5", "CloudTrail 암호화 설정",
     "aws cloudtrail describe-trails"),
    ("C-4.6", "CloudWatch 암호화 설정",
     "aws logs describe-log-groups"),
    ("C-4.7", "CloudTrail 로깅 설정",
     "aws cloudtrail describe-trails\naws cloudtrail get-trail-status --name <TRAIL_ARN>"),
    ("C-4.8", "인스턴스 로깅 설정",
     "aws ec2 describe-instances\n# CloudWatch 에이전트 설치 여부 수동 확인"),
    ("C-4.9", "RDS 로깅 설정",
     "aws rds describe-db-instances"),
    ("C-4.10", "S3 버킷 로깅 설정",
     "aws s3api get-bucket-logging --bucket <BUCKET>"),
    ("C-4.11", "VPC 플로우 로깅 설정",
     "aws ec2 describe-flow-logs"),
    ("C-4.12", "로그 보관 기간 설정",
     "aws logs describe-log-groups"),
    ("C-4.13", "백업 사용 여부",
     "aws backup list-backup-plans"),
    ("C-4.14", "AWS Config 서비스 활성화",
     "aws configservice describe-configuration-recorder-status"),
    ("C-4.15", "GuardDuty 활성화",
     "aws guardduty list-detectors\naws guardduty get-detector --detector-id <ID>"),
]


# ══════════════════════════════════════════════════════════════════
# 주통기 2026 가이드 명령어  (SK와 다른 항목만 변경, 나머지 동일)
# ══════════════════════════════════════════════════════════════════

_JTK_LINUX = [
    ("U-01", "root 계정 원격 접속 제한",
     "cat /etc/ssh/sshd_config | grep PermitRootLogin\ncat /etc/securetty"),
    ("U-02", "패스워드 복잡도 설정 (정책 기반)",
     "cat /etc/security/pwquality.conf\ncat /etc/pam.d/common-password\ncat /etc/pam.d/system-auth"),
    ("U-03", "계정 잠금 임계값 설정",
     "cat /etc/pam.d/common-auth\ncat /etc/pam.d/system-auth"),
    ("U-04", "패스워드 파일 보호",
     "ls -al /etc/passwd\nls -al /etc/shadow"),
    ("U-05", "root 이외 UID 0 금지",
     "awk -F: '$3 == 0 { print $0 }' /etc/passwd"),
    ("U-06", "root 계정 su 제한",
     "cat /etc/pam.d/su\ncat /etc/group | grep wheel"),
    ("U-07", "패스워드 최소 길이 설정",
     "grep PASS_MIN_LEN /etc/login.defs\ngrep minlen /etc/security/pwquality.conf"),
    ("U-08", "패스워드 최대 사용 기간 설정",
     "grep PASS_MAX_DAYS /etc/login.defs"),
    ("U-09", "패스워드 최소 사용 기간 설정",
     "grep PASS_MIN_DAYS /etc/login.defs"),
    ("U-10", "불필요한 계정 제거",
     'egrep "^(lp|uucp|nuucp):" /etc/passwd | egrep -v "false|nologin"'),
    ("U-11", "관리자 그룹 최소 계정 포함",
     "cat /etc/group | grep -E '^wheel|^sudo'"),
    ("U-12", "계정 없는 GID 금지",
     "cat /etc/passwd\ncat /etc/group"),
    ("U-13", "동일 UID 금지 / SHA-512 암호화 확인 (2026 신규)",
     "awk -F: '{print $3}' /etc/passwd | sort | uniq -d\ngrep -E 'SHA|sha' /etc/login.defs /etc/pam.d/common-password 2>/dev/null"),
    ("U-14", "사용자 shell 점검",
     r'grep -v "nologin\|false" /etc/passwd | awk -F: \'$7!="" {print $1,$7}\''),
    ("U-15", "세션 타임아웃 설정",
     "cat /etc/profile | grep TMOUT\ncat /etc/bashrc | grep TMOUT"),
    ("U-16", "홈 디렉터리 소유자·권한 설정",
     "ls -al /home"),
    ("U-17", "시스템 시작 스크립트 권한 설정 (2026 신규)",
     "ls -al /etc/rc*.d/* 2>/dev/null\nls -al /etc/inittab 2>/dev/null"),
    ("U-18", "world writable 파일 점검",
     "find / -type f -perm -o+w -exec ls -l {} \\; 2>/dev/null"),
    ("U-19", "/dev 불필요 파일 존재",
     "find /dev -type f -exec ls -l {} \\; 2>/dev/null"),
    ("U-20", "SetUID/SetGID 파일 점검",
     r"find / -user root -type f \( -perm -4000 -o -perm -2000 \) -exec ls -al {} \; 2>/dev/null"),
    ("U-22", "/etc/passwd 소유자·권한",
     "ls -al /etc/passwd"),
    ("U-23", "/etc/shadow 소유자·권한",
     "ls -al /etc/shadow"),
    ("U-24", "/etc/hosts 소유자·권한",
     "ls -al /etc/hosts"),
    ("U-27", "/etc/services 소유자·권한",
     "ls -al /etc/services"),
    ("U-29", "최신 보안 패치 적용",
     "uname -a\ncat /etc/os-release"),
    ("U-30", "ftpusers 파일 소유자·권한",
     "ls -al /etc/ftpusers 2>/dev/null"),
    ("U-31", "ftp root 계정 제한",
     "cat /etc/ftpusers 2>/dev/null"),
    ("U-32", "at 파일 소유자·권한",
     "ls -al /etc/at.allow 2>/dev/null\nls -al /etc/at.deny 2>/dev/null"),
    ("U-33", "cron 파일 소유자·권한",
     "ls -al /etc/crontab 2>/dev/null\nls -al /etc/cron.d/ 2>/dev/null"),
    ("U-34", "hosts.equiv 파일 제거",
     "ls -al /etc/hosts.equiv 2>/dev/null"),
    ("U-35", "공유 서비스 전체 익명 접근 제한 (2026 확장)",
     "cat /etc/vsftpd.conf 2>/dev/null | grep anonymous\ncat /etc/samba/smb.conf 2>/dev/null | grep 'guest ok'"),
    ("U-36", "r계열 서비스 비활성화",
     "systemctl is-active rsh rlogin rexec 2>/dev/null"),
    ("U-37", "NFS 설정 파일 점검",
     "cat /etc/exports 2>/dev/null"),
    ("U-42", "tftp, talk 서비스 비활성화",
     "systemctl is-active tftp talk 2>/dev/null"),
    ("U-46", "DNS 보안 버전 패치",
     "named -v 2>/dev/null"),
    ("U-47", "DNS Zone Transfer 설정",
     "cat /etc/named.conf 2>/dev/null | grep allow-transfer"),
    ("U-48", "Apache 디렉터리 리스팅 제거",
     "cat /etc/apache2/apache2.conf 2>/dev/null | grep -i options"),
    ("U-51", "DNS 취약한 동적 업데이트 설정 금지 (2026 신규)",
     "cat /etc/named.conf 2>/dev/null | grep allow-update"),
    ("U-52", "Telnet 서비스 비활성화 (2026 신규)",
     "systemctl is-active telnet 2>/dev/null\nss -tlnup | grep :23"),
    ("U-53", "FTP 서비스 정보 노출 방지 (2026 신규)",
     "cat /etc/vsftpd.conf 2>/dev/null | grep -E 'banner|ftpd_banner'"),
    ("U-54", "FTP 암호화 설정 (2026 신규)",
     "cat /etc/vsftpd.conf 2>/dev/null | grep -E 'ssl|tls'"),
    ("U-55", "FTP 접근 제어 설정 (2026 신규)",
     "cat /etc/vsftpd.conf 2>/dev/null | grep -E 'userlist|chroot'"),
    ("U-59", "SNMP v3 설정 (2026 신규)",
     "cat /etc/snmp/snmpd.conf 2>/dev/null | grep -E 'createUser|rouser|rwuser'"),
    ("U-60", "SNMP Community String 복잡성",
     "cat /etc/snmp/snmpd.conf 2>/dev/null | grep -E '^rocommunity|^rwcommunity'"),
    ("U-61", "SNMP ACL 설정 (2026 신규)",
     "cat /etc/snmp/snmpd.conf 2>/dev/null"),
    ("U-62", "로그온 배너 설정",
     "cat /etc/issue 2>/dev/null\ncat /etc/issue.net 2>/dev/null"),
    ("U-63", "sudo 명령어 접근 관리 (2026 신규)",
     "cat /etc/sudoers 2>/dev/null | grep -v '^#'\nls -al /etc/sudoers.d/ 2>/dev/null"),
    ("U-64", "syslog 로그 설정",
     "cat /etc/rsyslog.conf 2>/dev/null"),
    ("U-65", "NTP 시각 동기화 (2026 신규)",
     "systemctl is-active ntpd chronyd 2>/dev/null\ntimedatectl status 2>/dev/null\ncat /etc/ntp.conf 2>/dev/null | grep server"),
    ("U-67", "로그 디렉터리 소유자·권한 설정 (2026 신규)",
     "ls -ald /var/log\nstat -c '%a %U' /var/log"),
]

_JTK_WINDOWS = _SK_WINDOWS + [
    ("W-1.9", "통합 계정 정책 점검 (2026 신규)",
     'net accounts\ncmd /c "secedit /export /cfg C:\\Windows\\Temp\\wss2.cfg /quiet && type C:\\Windows\\Temp\\wss2.cfg"'),
    ("W-3.5", "Windows Defender Firewall 활성화 (2026 신규)",
     "netsh advfirewall show allprofiles state"),
    ("W-3.6", "SMB v1 비활성화 (2026 신규)",
     'powershell -NoProfile -NonInteractive -Command "Get-SmbServerConfiguration | Select EnableSMB1Protocol"\nreg query "HKLM\\SYSTEM\\CurrentControlSet\\Services\\LanmanServer\\Parameters" /v SMB1'),
    ("W-4.4", "NTP 시간 동기화 설정 (2026 신규)",
     "sc query W32Time\nw32tm /query /status"),
    ("W-5.17", "이벤트 로그 크기 및 보존 설정 (2026 신규)",
     "wevtutil gl Security\nwevtutil gl System\nwevtutil gl Application"),
    ("W-5.18", "서비스 계정 권한 최소화 (2026 신규)",
     'powershell -NoProfile -NonInteractive -Command "Get-WmiObject Win32_Service | Where-Object {$_.StartName -notin @(\'LocalSystem\',\'NT AUTHORITY\\\\LocalService\',\'NT AUTHORITY\\\\NetworkService\')} | Select Name,StartName"'),
    ("W-8.3", "Windows Update 자동 업데이트 설정 (2026 신규)",
     'reg query "HKLM\\SOFTWARE\\Policies\\Microsoft\\Windows\\WindowsUpdate\\AU" /v AUOptions 2>nul'),
]

_JTK_DBMS = [
    ("D-01", "DBMS 포트 외부 노출 여부",
     "# MySQL\nnc -zv <host> 3306\n# PostgreSQL\nnc -zv <host> 5432\n# MSSQL\nnc -zv <host> 1433"),
    ("D-02", "불필요 계정 제거",
     "# MySQL\nSELECT user, host, authentication_string FROM mysql.user;\n# PostgreSQL\nSELECT usename, passwd FROM pg_shadow;\n# MSSQL\nSELECT name, is_disabled FROM sys.server_principals WHERE type = 'S';"),
    ("D-03", "비밀번호 사용기간·복잡도 설정",
     "# MySQL\nSHOW VARIABLES LIKE 'validate_password%';\n# MSSQL\nSELECT name, is_policy_checked FROM sys.sql_logins;"),
    ("D-04", "관리자 권한 최소화",
     "# MySQL\nSELECT user, host FROM mysql.user WHERE user='root' AND host != 'localhost';\n# MSSQL\nSELECT m.name FROM sys.server_role_members rm JOIN sys.server_principals m ON rm.member_principal_id=m.principal_id JOIN sys.server_principals r ON rm.role_principal_id=r.principal_id WHERE r.name='sysadmin';"),
    ("D-05", "감사 로깅 설정",
     "# MySQL\nSHOW VARIABLES LIKE 'general_log%';\n# PostgreSQL\nSHOW log_statement;\n# MSSQL\nSELECT * FROM sys.server_audits;"),
    ("D-06", "데이터 디렉토리 권한",
     "# MySQL\nSHOW VARIABLES LIKE 'datadir';\n# PostgreSQL\nSHOW data_directory;"),
    ("D-07", "root 권한으로 서비스 구동 제한 (2026 신규)",
     "# MySQL\ngrep -i '^user' /etc/mysql/my.cnf /etc/my.cnf 2>/dev/null\nps aux | grep mysqld\n# PostgreSQL\nps aux | grep postgres"),
    ("D-08", "안전한 암호화 알고리즘 사용 (2026 신규)",
     "# MySQL\nSELECT user, plugin FROM mysql.user;\n# PostgreSQL\nSHOW password_encryption;\n# MSSQL\nSELECT name, password_hash FROM sys.sql_logins;"),
    ("D-09", "계정 잠금 정책 설정",
     "# MySQL\nSHOW VARIABLES LIKE 'connection_control%';\n# MSSQL\nSELECT name, is_policy_checked FROM sys.sql_logins;"),
    ("D-12", "안전한 리스너 비밀번호 설정",
     "# Oracle 전용 항목 — MySQL/PostgreSQL/MSSQL 해당 없음 (N/A)"),
    ("D-23", "xp_cmdshell 사용 제한 — MSSQL 전용 (2026 신규)",
     "SELECT value_in_use FROM sys.configurations WHERE name = 'xp_cmdshell';"),
    ("D-24", "Registry Procedure 권한 제한 — MSSQL 전용 (2026 신규)",
     "SELECT grantee_principal_id, permission_name FROM sys.database_permissions WHERE class_desc='OBJECT' AND major_id=OBJECT_ID('xp_regread');"),
]

_JTK_ORACLE = [
    ("D-01", "DBMS 포트 외부 노출 여부",
     "nc -zv <RDS_ENDPOINT> 1521"),
    ("D-02", "불필요 계정 제거",
     "SELECT username, account_status, password_versions FROM dba_users ORDER BY username;"),
    ("D-03", "비밀번호 사용기간·복잡도 설정",
     "SELECT resource_name, limit FROM dba_profiles WHERE profile='DEFAULT' AND resource_type='PASSWORD';"),
    ("D-04", "관리자 권한 최소화",
     "SELECT grantee, granted_role FROM dba_role_privs WHERE granted_role IN ('DBA','SYSDBA','SYSOPER');\nSELECT grantee, privilege FROM dba_sys_privs WHERE privilege IN ('CREATE USER','DROP USER','ALTER USER','GRANT ANY PRIVILEGE');"),
    ("D-05", "감사 로깅 설정",
     "SELECT value FROM v$parameter WHERE name='audit_trail';\nSELECT value FROM v$parameter WHERE name='unified_audit_trail';"),
    ("D-06", "데이터 디렉토리 권한",
     "SELECT value FROM v$parameter WHERE name='db_create_file_dest';\n# RDS: OS 직접 접근 불가 (N/A)"),
    ("D-07", "root 권한으로 서비스 구동 제한 (2026 신규)",
     "# RDS: 관리형 서비스 — OS 접근 불가 (N/A)"),
    ("D-08", "안전한 암호화 알고리즘 사용 (2026 신규)",
     "SELECT username, password_versions FROM dba_users;"),
    ("D-09", "계정 잠금 정책 설정",
     "SELECT resource_name, limit FROM dba_profiles WHERE profile='DEFAULT' AND resource_name IN ('FAILED_LOGIN_ATTEMPTS','PASSWORD_LOCK_TIME');"),
    ("D-10", "불필요한 PL/SQL 패키지 권한 제한",
     "SELECT grantee, table_name, privilege FROM dba_tab_privs WHERE table_name IN ('UTL_SMTP','UTL_TCP','UTL_HTTP','UTL_FILE','DBMS_SQL','DBMS_JOB') AND grantee='PUBLIC';"),
    ("D-11", "데이터베이스 링크 관리",
     "SELECT owner, db_link, username, host FROM dba_db_links;"),
    ("D-12", "안전한 리스너 비밀번호 설정",
     "lsnrctl status\ncat $ORACLE_HOME/network/admin/listener.ora"),
    ("D-13", "트리거 관리",
     "SELECT owner, trigger_name, status FROM dba_triggers WHERE owner NOT IN ('SYS','SYSTEM');"),
]

_JTK_AWS = [
    ("C-1.1", "IAM 사용자 계정 관리",
     "aws iam list-users\naws iam list-attached-user-policies --user-name <USER>\naws iam list-groups-for-user --user-name <USER>"),
    ("C-1.2", "IAM 사용자 계정 단일화",
     "aws iam list-users"),
    ("C-1.3", "IAM 사용자 계정 식별",
     "aws iam list-user-tags --user-name <USER>"),
    ("C-1.4", "IAM 그룹 사용자 관리",
     "aws iam list-groups\naws iam get-group --group-name <GROUP>"),
    ("C-1.9", "MFA 설정",
     "aws iam get-account-summary\naws iam list-mfa-devices --user-name <USER>\naws iam list-virtual-mfa-devices"),
    ("C-1.10", "AWS 계정 패스워드 정책",
     "aws iam get-account-password-policy"),
    ("C-1.11", "IAM 액세스 키 교체 주기 점검 (2026 신규)",
     "aws iam list-users\naws iam list-access-keys --user-name <USER>"),
    ("C-2.1", "인스턴스 서비스 정책 관리",
     "# 수동 점검\naws iam list-roles"),
    ("C-2.2", "네트워크 서비스 정책 관리",
     "# 수동 점검\naws iam list-policies --scope Local"),
    ("C-2.3", "기타 서비스 정책 관리",
     "# 수동 점검\naws iam list-policies --scope Local"),
    ("C-3.1", "VPC 네트워크 서브넷 관리",
     "aws ec2 describe-security-groups\naws ec2 describe-network-acls\naws ec2 describe-internet-gateways\naws ec2 describe-nat-gateways"),
    ("C-3.2", "가상 네트워크 리소스 관리",
     "aws ec2 describe-route-tables"),
    ("C-3.3", "접근 제어 설정 관리",
     "aws ec2 describe-security-groups"),
    ("C-3.4", "스토리지 리소스 퍼블릭 접근 관리",
     "aws s3api list-buckets\naws s3api get-public-access-block --bucket <BUCKET>\naws rds describe-db-subnet-groups"),
    ("C-4.1", "관계형 데이터베이스 암호화 설정",
     "aws rds describe-db-instances"),
    ("C-4.2", "통신구간 암호화 설정",
     "# 수동 점검 — TLS/SSL, VPN, SSH 적용 여부 확인"),
    ("C-4.3", "클라우드 서비스 사용자 계정 로깅 설정",
     "aws cloudtrail describe-trails\naws cloudtrail get-trail-status --name <TRAIL_ARN>"),
    ("C-4.4", "인스턴스 로깅 설정",
     "aws ec2 describe-instances\n# CloudWatch 에이전트 설치 여부 수동 확인"),
    ("C-4.5", "관계형 데이터베이스 로깅 설정",
     "aws rds describe-db-instances"),
    ("C-4.6", "오브젝트 스토리지 버킷 로깅 설정",
     "aws s3api list-buckets\naws s3api get-bucket-logging --bucket <BUCKET>"),
    ("C-4.7", "로그 보관 기간 설정",
     "aws logs describe-log-groups"),
    ("C-4.8", "백업 사용 여부",
     "aws backup list-backup-plans"),
    ("C-4.9", "AWS Config 서비스 활성화 (2026 신규)",
     "aws configservice describe-configuration-recorder-status"),
    ("C-4.10", "GuardDuty 활성화 (2026 신규)",
     "aws guardduty list-detectors\naws guardduty get-detector --detector-id <ID>"),
]


# ══════════════════════════════════════════════════════════════════
# 가이드별 도메인 매핑
# ══════════════════════════════════════════════════════════════════

_CONN_COMMANDS = [
    ("CONN-01", "SSH 접속 (PEM 키 인증)",
     "ssh -i /path/to/key.pem -p 22 ec2-user@<HOST>"),
    ("CONN-02", "SSH 접속 (패스워드 인증)",
     "ssh -p 22 user@<HOST>"),
    ("CONN-03", "SSH 접속 (알려진 호스트 검사 비활성화 — 테스트용)",
     "ssh -o StrictHostKeyChecking=no -i /path/to/key.pem ec2-user@<HOST>"),
    ("CONN-04", "AWS SSM 세션 시작",
     "aws ssm start-session \\\n"
     "    --region ap-northeast-2 \\\n"
     "    --target i-xxxxxxxxxxxxxxxxx"),
    ("CONN-05", "AWS SSM — Oracle RDS 포트 포워딩",
     "aws ssm start-session \\\n"
     "    --region ap-northeast-2 \\\n"
     "    --target i-xxxxxxxxxxxxxxxxx \\\n"
     "    --document-name AWS-StartPortForwardingSessionToRemoteHost \\\n"
     '    --parameters \'{"host":["your-rds.rds.amazonaws.com"],'
     '"portNumber":["1521"],"localPortNumber":["11521"]}\''),
    ("CONN-06", "AWS SSM — 임시 자격증명 사용",
     "export AWS_ACCESS_KEY_ID=AKIA...\n"
     "export AWS_SECRET_ACCESS_KEY=xxxx\n"
     "export AWS_SESSION_TOKEN=xxxx\n"
     "aws ssm start-session --region ap-northeast-2 --target i-xxxxxxxxxxxxxxxxx"),
    ("CONN-07", "Docker 컨테이너 접속 확인",
     "docker ps\ndocker exec -it <CONTAINER_ID> /bin/bash"),
    ("CONN-08", "AWS CLI 자격증명 확인",
     "aws sts get-caller-identity"),
    ("CONN-09", "SSM 연결 가능 인스턴스 목록 조회",
     "aws ssm describe-instance-information \\\n"
     "    --region ap-northeast-2 \\\n"
     '    --query "InstanceInformationList[*].[InstanceId,PingStatus,PlatformType]" \\\n'
     "    --output table"),
    ("CONN-10", "session-manager-plugin 설치 확인",
     "session-manager-plugin --version\n"
     "# 미설치 시:\n"
     "# macOS:   brew install --cask session-manager-plugin\n"
     "# Linux:   https://docs.aws.amazon.com/systems-manager/latest/userguide/session-manager-working-with-install-plugin.html"),
]

GUIDE_COMMANDS = {
    "sk": {
        "OS - Linux":            _SK_LINUX,
        "OS - Windows":          _SK_WINDOWS,
        "DBMS (MySQL/PG/MSSQL)": _SK_DBMS,
        "Oracle":                _SK_ORACLE,
        "AWS Cloud":             _SK_AWS,
        "연결 설정 (SSH/SSM)":   _CONN_COMMANDS,
    },
    "jtk": {
        "OS - Linux":            _JTK_LINUX,
        "OS - Windows":          _JTK_WINDOWS,
        "DBMS (MySQL/PG/MSSQL)": _JTK_DBMS,
        "Oracle":                _JTK_ORACLE,
        "AWS Cloud":             _JTK_AWS,
        "연결 설정 (SSH/SSM)":   _CONN_COMMANDS,
    },
}


# ══════════════════════════════════════════════════════════════════
# 출력 함수
# ══════════════════════════════════════════════════════════════════

def export_text(output_path: str, guide: str = "sk", domains: list | None = None):
    data = GUIDE_COMMANDS.get(guide, GUIDE_COMMANDS["sk"])
    guide_label = "SK Shieldus 표준 보안 가이드" if guide == "sk" else "주요정보통신기반시설 기술적 취약점 분석평가 가이드 (2026)"

    lines = [f"보안 취약점 진단 명령어 목록", f"기준 가이드: {guide_label}", "=" * 70, ""]

    for domain, items in data.items():
        if domains and domain not in domains:
            continue
        lines += [f"{'=' * 70}", f"  {domain}", f"{'=' * 70}"]
        for cid, name, cmd in items:
            lines += [f"\n[{cid}] {name}", "-" * 50]
            lines += [f"  {c}" for c in cmd.strip().splitlines()]
        lines.append("")

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def export_excel(output_path: str, guide: str = "sk", domains: list | None = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = GUIDE_COMMANDS.get(guide, GUIDE_COMMANDS["sk"])
    guide_label = "SK Shieldus 표준" if guide == "sk" else "주통기 2026"

    wb = Workbook()
    ws = wb.active
    ws.title = f"진단 명령어 ({guide_label})"

    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(fill_type="solid", fgColor="1F4E79")
    dfill = PatternFill(fill_type="solid", fgColor="2E75B6")
    dfont = Font(bold=True, color="FFFFFF", size=10)
    center  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin   = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 70

    for col, h in enumerate(["도메인", "항목 ID", "항목명", "명령어"], 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = center; cell.border = border
    ws.row_dimensions[1].height = 22

    row = 2
    for domain, items in data.items():
        if domains and domain not in domains:
            continue
        start_row = row
        for cid, name, cmd in items:
            ws.cell(row=row, column=1, value=domain).alignment = center
            ws.cell(row=row, column=2, value=cid).alignment = center
            ws.cell(row=row, column=3, value=name).alignment = left_top
            ws.cell(row=row, column=4, value=cmd).alignment = left_top
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                if col == 1:
                    cell.font = dfont; cell.fill = dfill
            ws.row_dimensions[row].height = max(cmd.count("\n") + 1, 1) * 15
            row += 1
        if row - 1 > start_row:
            ws.merge_cells(f"A{start_row}:A{row - 1}")

    ws.freeze_panes = "A2"
    wb.save(output_path)


# ══════════════════════════════════════════════════════════════════
# 단독 실행 (CLI)
# ══════════════════════════════════════════════════════════════════

def main():
    print("\n보안 진단 명령어 추출 도구")
    print("=" * 40)
    print("가이드 선택:")
    print("  1. SK Shieldus 표준 보안 가이드")
    print("  2. 주통기 2026 가이드")
    gc = input("\n선택 (1/2): ").strip()
    guide = "jtk" if gc == "2" else "sk"

    data = GUIDE_COMMANDS[guide]
    domain_list = list(data.keys())
    print("\n도메인 선택 (쉼표로 구분, 전체는 Enter):")
    for i, d in enumerate(domain_list, 1):
        print(f"  {i}. {d}")
    dc = input("\n선택 (예: 1,3 또는 Enter=전체): ").strip()
    if dc:
        idxs = [int(x.strip()) - 1 for x in dc.split(",") if x.strip().isdigit()]
        domains = [domain_list[i] for i in idxs if 0 <= i < len(domain_list)]
    else:
        domains = None

    print("\n출력 형식:")
    print("  1. Excel (.xlsx)")
    print("  2. 텍스트 (.txt)")
    print("  3. 둘 다")
    fc = input("\n선택 (1/2/3): ").strip()

    guide_tag = "주통기2026" if guide == "jtk" else "SK표준"
    base = f"진단_명령어_{guide_tag}"

    if fc in ("1", "3"):
        export_excel(f"{base}.xlsx", guide, domains)
        print(f"[완료] Excel 저장: {base}.xlsx")
    if fc in ("2", "3"):
        export_text(f"{base}.txt", guide, domains)
        print(f"[완료] 텍스트 저장: {base}.txt")


if __name__ == "__main__":
    main()
